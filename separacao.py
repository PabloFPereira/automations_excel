import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# Função para criar diretórios se não existirem
def ensure_dir(directory):
    if not os.path.exists(directory):
        os.makedirs(directory)

# Carregar o arquivo Excel
file_path = r'Neurociencia\Manutenções de Painel - Neurociencia.xlsx'  # Substitua pelo caminho do seu arquivo
df = pd.read_excel(file_path)

# Iterar por cada representante/setor único na coluna SETOR
for setor in df['SETOR'].unique():
    # Filtrar o DataFrame para o setor/representante atual
    df_setor = df[df['SETOR'] == setor]
    
    # Definir o nome do arquivo de saída
    output_dir = r'Neurociencia\Arquivos Separados'  # Substitua pelo caminho desejado para salvar os arquivos
    ensure_dir(output_dir)
    output_file = os.path.join(output_dir, f'{setor}.xlsx')
    
    # Salvar o DataFrame filtrado em um novo arquivo Excel
    df_setor.to_excel(output_file, index=False)
    
    # Carregar o arquivo salvo para aplicar formatação
    wb = load_workbook(output_file)
    ws = wb.active
    
    # Aplicar cor de fundo preto e fonte branca na primeira linha
    header_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    
    # Congelar a primeira linha
    ws.freeze_panes = 'A2'
    
    # Adicionar filtro à primeira linha
    ws.auto_filter.ref = ws.dimensions
    
    # Ajustar a largura das colunas após aplicar o filtro
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter # Coluna A, B, C, etc.
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width
    
    # Adicionar lista de opções à coluna L
    dv = DataValidation(type="list", formula1='"Incluir,Excluir,Editar Informações"', showDropDown=True)
    ws.add_data_validation(dv)
    dv.add(f'L2:L{ws.max_row}')
    
    # Salvar o arquivo com formatação
    wb.save(output_file)

print("Arquivos separados e formatados foram criados com sucesso.")
