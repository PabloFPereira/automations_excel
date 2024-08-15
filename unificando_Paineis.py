import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

def formatar_arquivo(arquivo_saida):
    # Carrega o workbook e aplica formatação
    wb = load_workbook(arquivo_saida)
    
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        # Aplica a cor de fundo preta e fonte branca na primeira linha
        for cell in ws[1]:
            cell.font = Font(color="FFFFFF", bold=True)
            cell.fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
        # Adiciona autofiltro
        ws.auto_filter.ref = ws.dimensions
    
    wb.save(arquivo_saida)

def unificar_abas(pasta, equipe, arquivo_saida):
    # Dicionários para armazenar dataframes das abas
    abas_instituicoes = []
    abas_profissionais = []
    dataframes = {}
    
    # Percorre todos os arquivos na pasta
    for nome_arquivo in os.listdir(pasta):
        caminho_arquivo = os.path.join(pasta, nome_arquivo)
        # Verifica se é um arquivo Excel
        if os.path.isfile(caminho_arquivo) and nome_arquivo.endswith('.xlsx'):
            # Lê o arquivo Excel
            xls = pd.ExcelFile(caminho_arquivo)
            if equipe in ["Specialty Care", "Hospitalar"]:
                # Lê as abas Instituições e Profissionais separadamente
                if "Instituições" in xls.sheet_names:
                    df_instituicoes = pd.read_excel(caminho_arquivo, sheet_name="Instituições")
                    abas_instituicoes.append(df_instituicoes)
                if "Profissionais" in xls.sheet_names:
                    df_profissionais = pd.read_excel(caminho_arquivo, sheet_name="Profissionais")
                    abas_profissionais.append(df_profissionais)
                print(f'Arquivo {nome_arquivo} foi lido e as abas Instituições e Profissionais foram adicionadas à lista')
            else:
                # Para outras equipes, lê todas as abas
                for sheet in xls.sheet_names:
                    df = pd.read_excel(caminho_arquivo, sheet_name=sheet)
                    if sheet not in dataframes:
                        dataframes[sheet] = []
                    dataframes[sheet].append(df)
                print(f'Arquivo {nome_arquivo} foi lido e todas as abas foram adicionadas à lista')

    # Cria um ExcelWriter para salvar as abas no mesmo arquivo
    with pd.ExcelWriter(f"{arquivo_saida}.xlsx", engine='openpyxl') as writer:
        if equipe in ["Specialty Care", "Hospitalar"]:
            # Concatena e salva as abas separadamente
            if abas_instituicoes:
                df_instituicoes_unificado = pd.concat(abas_instituicoes, ignore_index=True)
                df_instituicoes_unificado.to_excel(writer, sheet_name='Instituições', index=False)
            if abas_profissionais:
                df_profissionais_unificado = pd.concat(abas_profissionais, ignore_index=True)
                df_profissionais_unificado.to_excel(writer, sheet_name='Profissionais', index=False)
        else:
            # Concatena e salva todas as abas
            for sheet, dfs in dataframes.items():
                df_unificado = pd.concat(dfs, ignore_index=True)
                df_unificado.to_excel(writer, sheet_name=sheet, index=False)
    
    print(f'Todos os arquivos foram unificados para a equipe {equipe}')

    # Formata o arquivo de saída
    formatar_arquivo(f"{arquivo_saida}.xlsx")

def selecionar_equipe():
    equipes = ["Specialty Care", "Raras", "Neurociências", "Hospitalar"]
    print("Selecione a equipe:")
    for i, equipe in enumerate(equipes):
        print(f"{i + 1}. {equipe}")
    
    escolha = int(input("Digite o número correspondente à equipe: "))
    return equipes[escolha - 1]

# Exemplo de uso
pasta = 'C:/Users/PPEREIRA01/Documents/Python/Unificando Paineis/Arquivos Separador por Setor'  # Substitua pelo caminho da sua pasta
equipe = selecionar_equipe()
arquivo_saida = f'C:/Users/PPEREIRA01/Documents/Python/Unificando Paineis/Manutenções de Painel - {equipe}'  # Substitua pelo caminho do arquivo de saída

unificar_abas(pasta, equipe, arquivo_saida)
